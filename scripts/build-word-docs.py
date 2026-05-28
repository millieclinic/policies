#!/usr/bin/env python3
"""build-word-docs.py — assemble the exec-facing bundle of Word/Excel/PDF files.

Strategy (maximum fidelity to the originals):
  * For policies that are unchanged from `1. original_docs_markdown/`, copy the
    *original* source file (`.docx`, `.xlsx`, `.pdf`) directly from
    `Policy Docs/`. This preserves all of the original Word formatting — fonts,
    headings, numbered lists, tables — exactly as drafted.
  * For policies that were modified in `3. final_policies/` (e.g., the 3 short
    ECH-required clauses appended), pandoc-convert the modified Markdown to
    `.docx`. These will not look identical to the originals, but they'll be
    readable Word docs.
  * `CHANGES.md` (and any other generated Markdown without a source) is also
    pandoc-converted.

Re-run any time the final policies change. Auto-installs pandoc via Homebrew
on first run if missing.

Output: `policy_docs_word/`
"""

from __future__ import annotations

import filecmp
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
POLICY_DOCS = REPO / "1. Original Docs (Word)"
FOLDER_1 = POLICY_DOCS / "1.a Original Converted to MD"
FOLDER_2 = REPO / "2. Gaps" / "2.a Gaps Markdown"
OUT_GAP = REPO / "2. Gaps"
OUT = REPO / "3. Final Word"
FOLDER_3 = OUT / "3.a Final Markdown"


def ensure_pandoc() -> None:
    if shutil.which("pandoc") is None:
        print("Pandoc not found; installing via Homebrew…")
        if shutil.which("brew") is None:
            print(
                "ERROR: Homebrew not found. Install pandoc manually: "
                "https://pandoc.org/installing.html",
                file=sys.stderr,
            )
            sys.exit(1)
        subprocess.run(["brew", "install", "pandoc"], check=True)


def find_source(stem: str) -> Path | None:
    """Find an original file in Policy Docs/ that matches the given stem."""
    for ext in (".docx", ".pdf", ".xlsx", ".doc"):
        cand = POLICY_DOCS / f"{stem}{ext}"
        if cand.exists():
            return cand
    return None


def find_source_by_content(md_path: Path) -> Path | None:
    """Look for a folder-1 .md with byte-identical content, then return its
    matching `.docx` / `.pdf` / `.xlsx` source from POLICY_DOCS.

    This handles the case where folder-3 files were renamed (so name-match
    fails) but their content is still unchanged from the baseline — we want
    to ship the original Word file with perfect fidelity rather than
    pandoc-rendering it.
    """
    target_bytes = md_path.read_bytes()
    for f1_md in FOLDER_1.glob("*.md"):
        if f1_md.read_bytes() == target_bytes:
            return find_source(f1_md.stem)
    return None


_ROMAN = ["", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"]
_LETTERS = "abcdefghijklmnopqrstuvwxyz"


def rebuild_heading_hierarchy(text: str) -> str:
    """Restore the heading hierarchy that markitdown flattened into a numbered list.

    Some Millie policies came out of markitdown as a flat numbered list:

        1. **Scope**
        [paragraph]
        1. **Policy**
        2. It is Company's policy...
        3. This Policy applies...
        4. **Procedure**
        5. *Physical Safeguards*.
        6. Users must always lock...

    The original .docx had I/II/III headings, a/b/c sub-headings, and
    i/ii/iii items. Pandoc renders the flat list literally and loses all
    structure. This pre-pass detects the heading-like numbered items and
    converts them back to proper markdown headings so pandoc produces
    Word Heading 1 / Heading 2 styles.

    Patterns:
      "<num>. **Section**"        -> "## I. Section"   (Roman, h2_counter)
      "<num>. *Subsection*[.]"    -> "### a. Subsection" (letter, h3 per section)
      "<num>. <body text>"        -> "- <body text>"   (bullet under nearest heading)
    """
    lines = text.split("\n")

    # Only apply when the flat pattern is actually present — Framework-style
    # files that already use `# 1. Overview` markdown headings should pass through
    # unchanged.
    if not any(
        re.match(r"^\d+\. \*\*[^*]+\*\*\s*$", ln) for ln in lines
    ):
        return text

    out: list[str] = []
    h2_counter = 0
    h3_counter = 0
    for line in lines:
        # Top-level section heading: "1. **Scope**"
        m = re.match(r"^\d+\. \*\*([^*]+)\*\*\s*$", line)
        if m:
            h2_counter += 1
            h3_counter = 0
            num = _ROMAN[h2_counter] if h2_counter < len(_ROMAN) else str(h2_counter)
            out.append(f"## {num}. {m.group(1).strip()}")
            continue
        # Sub-heading: "5. *Physical Safeguards*."
        m = re.match(r"^\d+\. \*([^*]+)\*\.?\s*$", line)
        if m:
            h3_counter += 1
            letter = (
                _LETTERS[h3_counter - 1]
                if h3_counter - 1 < len(_LETTERS)
                else str(h3_counter)
            )
            out.append(f"### {letter}. {m.group(1).strip()}")
            continue
        # Numbered body item that follows a heading — render as bullet so it
        # nests visually under the heading rather than starting a new numbered
        # list at "1." (pandoc default).
        m = re.match(r"^\d+\. (.+)$", line)
        if m and h2_counter > 0:
            out.append(f"- {m.group(1).strip()}")
            continue
        out.append(line)
    return "\n".join(out)


def preprocess_md(src: Path) -> Path:
    """Strip docx→md conversion artifacts so pandoc → .docx renders cleanly."""
    text = src.read_text(encoding="utf-8")

    # Remove the 1-cell base64 logo table at the top: 2 lines like
    #   | **![](data:image/jpeg;base64...)** | |
    #   | --- | --- |
    text = re.sub(
        r"^\| \*\*!\[\]\(data:image[^)]*\)\*\* \| \|\n\| --- \| --- \|\n",
        "",
        text,
        flags=re.MULTILINE,
    )
    # Also handle png variants and the un-bolded form
    text = re.sub(
        r"^\| !\[\]\(data:image[^)]*\) \| \|\n\| --- \| --- \|\n",
        "",
        text,
        flags=re.MULTILINE,
    )

    # Remove any remaining inline base64 image placeholders
    text = re.sub(r"!\[\]\(data:image[^)]*\)", "", text)

    # Fix the very common docx→md typo ("Companypersonnel" missing a space)
    text = text.replace("Companypersonnel", "Company personnel")

    # Restore I/II/III heading hierarchy that markitdown flattened
    text = rebuild_heading_hierarchy(text)

    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".md", delete=False, encoding="utf-8"
    )
    tmp.write(text)
    tmp.close()
    return Path(tmp.name)


def pandoc_to_docx(src_md: Path, dst_docx: Path) -> None:
    cleaned = preprocess_md(src_md)
    try:
        subprocess.run(
            [
                "pandoc",
                str(cleaned),
                "--from",
                "markdown",
                "--to",
                "docx",
                "--standalone",
                "-o",
                str(dst_docx),
            ],
            check=True,
        )
    finally:
        cleaned.unlink(missing_ok=True)


def main() -> None:
    ensure_pandoc()

    if not FOLDER_3.exists():
        print(f"ERROR: {FOLDER_3} not found.", file=sys.stderr)
        sys.exit(1)

    OUT.mkdir(exist_ok=True)
    OUT_GAP.mkdir(exist_ok=True)
    # Clean rebuild — remove anything we previously wrote
    for f in OUT.iterdir():
        if f.is_file():
            f.unlink()
    for f in OUT_GAP.iterdir():
        if f.is_file():
            f.unlink()

    stats = {
        "copied_docx": 0,
        "copied_other": 0,
        "pandoc_modified": 0,
        "pandoc_generated": 0,
        "gap_docx": 0,
        "gap_xlsx": 0,
    }

    for md_file in sorted(FOLDER_3.glob("*.md")):
        stem = md_file.stem

        # CHANGES.md / PER-FILE-CHANGELOG.md → always pandoc (no source)
        if stem in ("CHANGES", "PER-FILE-CHANGELOG"):
            dst = OUT / f"{stem}.docx"
            pandoc_to_docx(md_file, dst)
            stats["pandoc_generated"] += 1
            print(f"  [pandoc, no source]  {dst.name}")
            continue

        # Try to find the original source (matching by name first, then by content
        # since folder-3 files may have been renamed).
        source = find_source(stem)
        if source is None:
            source = find_source_by_content(md_file)

        if source is not None:
            # Content-identical to the original — copy the source file for perfect fidelity.
            # Output filename uses the new (folder-3) name with the source extension.
            dst = OUT / f"{stem}{source.suffix}"
            shutil.copy2(source, dst)
            if source.suffix == ".docx":
                stats["copied_docx"] += 1
            else:
                stats["copied_other"] += 1
            print(f"  [copy {source.suffix:>5}]     {dst.name}")
        else:
            # Modified, merged, or new — pandoc the .md so the additions are reflected.
            dst = OUT / f"{stem}.docx"
            pandoc_to_docx(md_file, dst)
            stats["pandoc_modified"] += 1
            print(f"  [pandoc, modified]   {dst.name}")

    # ---------- Gap analysis bundle ----------
    # Convert the markdown gap docs to .docx and the coverage CSV to a formatted .xlsx
    # so an exec can review the analysis alongside the policies.
    if FOLDER_2.exists():
        print()
        print(f"Gap analysis (writing to {OUT_GAP.relative_to(REPO)}/):")
        for md_file in sorted(FOLDER_2.glob("*.md")):
            dst = OUT_GAP / f"{md_file.stem}.docx"
            pandoc_to_docx(md_file, dst)
            stats["gap_docx"] += 1
            print(f"  [pandoc]   {dst.name}")
        for csv_file in sorted(FOLDER_2.glob("*.csv")):
            dst = OUT_GAP / f"{csv_file.stem}.xlsx"
            try:
                import csv as _csv
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, PatternFill
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = "Coverage Matrix"
                with open(csv_file, newline="", encoding="utf-8") as f:
                    rows = list(_csv.reader(f))
                for row in rows:
                    ws.append(row)
                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="305496")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                # Column widths — generous for text-heavy columns, narrower for short ones
                widths = {
                    "A": 18, "B": 28, "C": 8, "D": 50, "E": 36,
                    "F": 22, "G": 36, "H": 16, "I": 18, "J": 50,
                    "K": 36, "L": 12, "M": 36, "N": 30, "O": 50, "P": 10,
                }
                for col_letter, w in widths.items():
                    ws.column_dimensions[col_letter].width = w
                # Wrap text + top-align all data cells
                wrap_align = Alignment(wrap_text=True, vertical="top")
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = wrap_align
                # Freeze the header row
                ws.freeze_panes = "A2"
                wb.save(dst)
                stats["gap_xlsx"] += 1
                print(f"  [xlsx]     {dst.name}")
            except Exception as e:
                # Fallback: plain copy of the CSV
                shutil.copy2(csv_file, OUT_GAP / csv_file.name)
                stats["gap_xlsx"] += 1
                print(f"  [copy csv (xlsx failed: {e})] {csv_file.name}")

    total = sum(stats.values())
    print()
    print(f"Done. {total} files written:")
    print(f"  • {OUT.relative_to(REPO)}/ ........... {stats['copied_docx'] + stats['copied_other'] + stats['pandoc_modified'] + stats['pandoc_generated']} policy files")
    print(f"  • {OUT_GAP.relative_to(REPO)}/ ............................. {stats['gap_docx'] + stats['gap_xlsx']} gap-analysis files")
    print(f"  copied .docx originals:  {stats['copied_docx']}")
    print(f"  copied .xlsx/.pdf:       {stats['copied_other']}")
    print(f"  pandoc (modified .md):   {stats['pandoc_modified']}")
    print(f"  pandoc (generated .md):  {stats['pandoc_generated']}")
    print(f"  gap-analysis .docx:      {stats['gap_docx']}")
    print(f"  gap-analysis .xlsx:      {stats['gap_xlsx']}")
    print()
    print("Share with execs:")
    print("  • Word — open .docx files directly")
    print("  • Excel — open .xlsx files directly")
    print("  • Google Drive — drag the whole folder in; .docx/.xlsx/.pdf all")
    print("    auto-open in the matching Google app")


if __name__ == "__main__":
    main()
